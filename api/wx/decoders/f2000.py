"""
Decoder for F.2000 yearly rainfall register (xlsx).

Validates and ingests daily precipitation (mm) from the standard
F.2000 form layout. One station per file, one sheet per file.

LAYOUT
======
Row 1: Magic word - F. 2000
Row 2: blank
Row 3: "REGISTER OF RAINFALL RECORDED DURING {year}"
Row 4: "at {station name}"
Row 5: identifier
Row 6: blank
Row 7: column headers (Day | JAN. | FEB. | ...)
Row 8: sub-headers (Measure | ML | MM | ...)
Row 9+: data rows (days 1-31)
"""

import calendar
import datetime
import logging
import re
import time

import openpyxl
import pytz
from celery import shared_task

from tempestas_api import settings
from wx.decoders.insert_raw_data import insert
from wx.decoders.manual_data import find_station_by_name
from wx.models import Variable

logger = logging.getLogger('surface.f2000')
db_logger = logging.getLogger('db')
FORMAT = "F2000"

PRECIP_VARIABLE_ID = Variable.objects.get(symbol='PRECIP').id
DAILY_SECONDS = 86400
TRACE_RAIN_VALUE = -0.1
TOLERANCE = 0.01
RAIN_DAY_THRESHOLD = 0.25

MONTHS = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN',
          'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']

# Fixed row positions (1-indexed for openpyxl)
ROW_IDENTIFIER = 1    # "F. 2000"
ROW_YEAR_HEADER = 3   # "REGISTER OF RAINFALL RECORDED DURING {year}"
ROW_STATION = 4       # "at" + station name
ROW_STATION_ID = 5    # 8-character station identifier
ROW_COL_HEADERS = 7   # Day | JAN. | FEB. | ...
ROW_SUB_HEADERS = 8   # Measure | ML | MM | ... (ml layout only)
ROW_DATA_START_ML = 9  # Day 1 data when ml=True (has sub-header row)
ROW_DATA_START_MM = 8  # Day 1 data when ml=False (no sub-header row)

# MM column indices (1-indexed for openpyxl) depend on layout
# ml=True:  A=Day, B=JAN_ML, C=JAN_MM, D=FEB_ML, E=FEB_MM, ... X=DEC_ML, Y=DEC_MM, Z=Day
# ml=False: A=Day, B=JAN, C=FEB, ... M=DEC, N=Day
ML_COL = {month_num: 2 * month_num for month_num in range(1, 13)}
MM_COL_ML = {month_num: 2 * month_num + 1 for month_num in range(1, 13)}
MM_COL_MMONLY = {month_num: month_num + 1 for month_num in range(1, 13)}


class F2000ValidationError(Exception):
    """Raised when one or more validation rules fail."""
    def __init__(self, errors):
        self.errors = errors
        super().__init__(f"{len(errors)} validation error(s): {'; '.join(errors)}")


def validate_structure(ws):
    """
    Validate worksheet structure. Returns metadata or raises F2000ValidationError.

    Auto-detects ML/MM vs MM-only layout from cell A8.
    Collects all errors before raising.

    Args:
        ws: openpyxl Worksheet

    Returns:
        metadata dict: {
            'year': int,
            'station_name': str,
            'station_id': str,
            'ml': bool,
        }
    """
    errors = []
    year = None
    station_name = None
    station_id = None

    # Auto-detect layout: ML layout has 'Measure' in cell A8
    cell_a8 = ws.cell(ROW_SUB_HEADERS, 1).value
    ml = cell_a8 is not None and str(cell_a8).strip() == 'Measure'

    # Rule 1 (single ws per file) removed, we need to update the checks below.

    # Rule 2: Check magic word in row 1
    if ws.cell(ROW_IDENTIFIER, 1).value != 'F. 2000':
        errors.append("Rule 2: Row 1, Cell 1 must contain 'F. 2000'")

    # Rule 3: Row 3 contains year header (may be in a merged cell, not col 1)
    row3 = " ".join(str(c.value) for c in ws[ROW_YEAR_HEADER] if c.value is not None)
    row3 = row3.lower()
    _match = re.search(r'register of rainfall recorded during (\d{4})', row3)
    if not _match:
        errors.append("Rule 3: Row 3 must contain 'REGISTER OF RAINFALL RECORDED DURING {year}'")
    else:
        year = int(_match.group(1))

    # Rule 4: Row 4 contains station name
    row4 = " ".join(str(c.value) for c in ws[ROW_STATION] if c.value is not None)
    _match = re.match(r'at (.+)$', row4)
    if not _match:
        errors.append("Rule 4: Row 4 must contain 'at {station name}'")
    else:
        station_name = _match.group(1).strip()

    # Rule 5: Row 5 contains 8-character station identifier
    row5 = " ".join(str(c.value) for c in ws[ROW_STATION_ID] if c.value is not None).strip()
    if len(row5) != 8:
        errors.append(f"Rule 5: Row 5 must contain 8 character station identifier, found '{row5}'")
    else:
        station_id = row5

    # Rule 6: Validate column headers on row 7
    _column_names = [c.value for c in ws[ROW_COL_HEADERS]]
    # Strip trailing None values from the list
    while _column_names and _column_names[-1] is None:
        _column_names.pop()

    if ml:
        expected_headers = ['Day', 'JAN.', None, 'FEB.', None, 'MAR.', None,
                            'APR.', None, 'MAY', None, 'JUNE', None, 'JULY',
                            None, 'AUG.', None, 'SEPT.', None, 'OCT.', None,
                            'NOV.', None, 'DEC.', None, 'Day']
    else:
        expected_headers = ['Day', 'JAN.', 'FEB.', 'MAR.', 'APR.', 'MAY',
                            'JUNE', 'JULY', 'AUG.', 'SEPT.', 'OCT.', 'NOV.',
                            'DEC.', 'Day']

    if _column_names != expected_headers:
        errors.append(f"Rule 6: Column headers on row 7 do not match expected layout")

    # Rule 7: Validate ML/MM sub-headers on row 8 (ml layout only)
    if ml:
        sub_headers = [c.value for c in ws[ROW_SUB_HEADERS]]
        while sub_headers and sub_headers[-1] is None:
            sub_headers.pop()
        expected_subheaders = ['Measure'] + ['ML', 'MM'] * 12
        if sub_headers != expected_subheaders:
            errors.append(f"Rule 7: Sub-headers on row 8 do not match expected ML/MM layout")

    # Rule 8: Validate day column (rows should contain days 1-31 in sequence)
    data_start = ROW_DATA_START_ML if ml else ROW_DATA_START_MM
    for day in range(1, 32):
        row_idx = data_start + day - 1
        cell_value = ws.cell(row_idx, 1).value
        try:
            if int(cell_value) != day:
                errors.append(f"Rule 8: Expected day {day} at row {row_idx}, found '{cell_value}'")
        except (ValueError, TypeError):
            errors.append(f"Rule 8: Expected day {day} at row {row_idx}, found '{cell_value}'")

    # Rule 9: Footer structure - Total row follows day 31
    total_row = data_start + 31
    total_label = ws.cell(total_row, 1).value
    if total_label is None or str(total_label).strip().lower() != 'total':
        errors.append(f"Rule 9: Expected 'Total' at row {total_row}, found '{total_label}'")

    # Rule 10: "No. of Days" row follows Total
    ndays_row = total_row + 1
    ndays_label = ws.cell(ndays_row, 1).value
    if ndays_label is None or 'no' not in str(ndays_label).strip().lower():
        errors.append(f"Rule 10: Expected 'No. of Days' at row {ndays_row}, found '{ndays_label}'")


    return errors, {
        'year': year,
        'station_name': station_name,
        'station_id': station_id,
        'ml': ml,
    }


# Content Validation

def _read_numeric(ws, row, col):
    """Read a cell value as float, returning None if empty/non-numeric."""
    value = ws.cell(row, col).value
    if value is None or str(value).strip() == '':
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def validate_content(ws, metadata):
    """
    Validate data values. Returns list of errors (empty = valid).
    Only called if Phase 1 (structure) passed.
    """
    errors = []
    year = metadata['year']
    ml = metadata['ml']
    mm_col = MM_COL_ML if ml else MM_COL_MMONLY
    data_start = ROW_DATA_START_ML if ml else ROW_DATA_START_MM
    total_row = data_start + 31
    ndays_row = total_row + 1

    # Per-cell validation
    has_nonzero = False
    for month_num in range(1, 13):
        col = mm_col[month_num]
        last_day = calendar.monthrange(year, month_num)[1]
        month_name = MONTHS[month_num - 1]

        for day in range(1, 32):
            row = data_start + day - 1
            cell_value = ws.cell(row, col).value

            if day > last_day:
                if cell_value is not None and str(cell_value).strip() != '':
                    errors.append(f"Data found for invalid date {month_name} {day}")
                continue

            if cell_value is None or str(cell_value).strip() == '':
                continue

            try:
                mm_value = float(cell_value)
            except (ValueError, TypeError):
                errors.append(f"Non-numeric value '{cell_value}' at day {day}, {month_name}")
                continue

            if mm_value != 0:
                has_nonzero = True

    if not has_nonzero:
        errors.append("File appears to be a blank template — all values are zero or empty")

    # validate ML to MM conversion (ml layout only)
    if ml:
        for month_num in range(1, 13):
            mm_c = mm_col[month_num]
            ml_c = ML_COL[month_num]
            last_day = calendar.monthrange(year, month_num)[1]
            month_name = MONTHS[month_num - 1]

            # Derive ratio from first valid pair (should we use average ratio?)
            ratio = None
            for day in range(1, last_day + 1):
                row = data_start + day - 1
                ml_val = _read_numeric(ws, row, ml_c)
                mm_val = _read_numeric(ws, row, mm_c)
                if ml_val is not None and mm_val is not None and mm_val != 0:
                    ratio = ml_val / mm_val
                    break

            if ratio is None:
                continue

            # Verify all rows use the same ratio
            for day in range(1, last_day + 1):
                row = data_start + day - 1
                ml_val = _read_numeric(ws, row, ml_c)
                mm_val = _read_numeric(ws, row, mm_c)
                if ml_val is None or mm_val is None:
                    continue
                if ml_val == 0 and mm_val == 0:
                    continue
                if mm_val == 0:
                    continue
                expected_mm = ml_val / ratio
                if abs(expected_mm - mm_val) > TOLERANCE:
                    errors.append(
                        f"ML to MM conversion mismatch at day {day}, {month_name}: "
                        f"ML={ml_val}, MM={mm_val}, expected MM={expected_mm:.4f}"
                    )

    # Total row sum check
    for month_num in range(1, 13):
        col = mm_col[month_num]
        last_day = calendar.monthrange(year, month_num)[1]
        month_name = MONTHS[month_num - 1]

        calculated_sum = 0.0
        for day in range(1, last_day + 1):
            row = data_start + day - 1
            mm_val = _read_numeric(ws, row, col)
            if mm_val is not None and mm_val > 0:
                calculated_sum += mm_val

        total_val = _read_numeric(ws, total_row, col)
        if total_val is not None and abs(calculated_sum - total_val) > TOLERANCE:
            errors.append(
                f"Total mismatch for {month_name}: "
                f"expected {calculated_sum:.2f}, found {total_val:.2f}"
            )

    # No. of Days check
    for month_num in range(1, 13):
        col = mm_col[month_num]
        last_day = calendar.monthrange(year, month_num)[1]
        month_name = MONTHS[month_num - 1]

        counted_days = 0
        for day in range(1, last_day + 1):
            row = data_start + day - 1
            mm_val = _read_numeric(ws, row, col)
            if mm_val is not None and mm_val >= RAIN_DAY_THRESHOLD:
                counted_days += 1

        ndays_val = _read_numeric(ws, ndays_row, col)
        if ndays_val is not None and int(ndays_val) != counted_days:
            errors.append(
                f"No. of Days mismatch for {month_name}: "
                f"expected {counted_days}, found {int(ndays_val)}"
            )

    return errors


# Now function to parse the data
def parse_data(ws, metadata, station, utc_offset):
    """
    Extract daily MM readings into raw_data tuples.

    Args:
        ws: openpyxl worksheet
        metadata: dict from validate_structure
        station: Station model instance
        utc_offset: timezone offset in minutes

    Returns:
        list of 15-element tuples for insert()
    """
    reads = []
    year = metadata['year']
    ml = metadata['ml']
    mm_col = MM_COL_ML if ml else MM_COL_MMONLY
    data_start = ROW_DATA_START_ML if ml else ROW_DATA_START_MM
    datetime_offset = pytz.FixedOffset(utc_offset)

    for month_num in range(1, 13):
        col = mm_col[month_num]
        last_day = calendar.monthrange(year, month_num)[1]

        for day in range(1, 32):
            if day > last_day:
                continue

            row = data_start + day - 1
            cell_value = ws.cell(row, col).value

            if cell_value is None or str(cell_value).strip() == '':
                measurement = settings.MISSING_VALUE
            else:
                measurement = float(cell_value)

            obs_date = datetime_offset.localize(datetime.datetime(year, month_num, day))

            reads.append((
                station.id,
                PRECIP_VARIABLE_ID,
                DAILY_SECONDS,
                obs_date,
                measurement,
                None, None, None, None, None,
                None, None, None, None,
                True,
            ))

    return reads


# ─── Entry Point ────────────────────────────────────────────────

@shared_task
def read_file(filename, highfrequency_data=False, station_object=None, utc_offset=settings.TIMEZONE_OFFSET, override_data_on_conflict=False):
    """Read an F.2000 yearly rainfall register and ingest daily precipitation (mm)."""

    logger.info(f'processing {filename}')
    start = time.time()
    reads = []

    # Load workbook
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
    except FileNotFoundError as fnf:
        logger.error(repr(fnf))
        print('No such file or directory {}.'.format(filename))
        raise
    except Exception as e:
        logger.error(repr(e))
        raise

    # iterate over worksheets, skipping bad sheets like hydro decoder
    for sheet in wb.sheetnames:
        if sheet in ('F2000 Template (ml)','Sheet 1 (mm)'):  # skip template sheets
            continue
        ws = wb[sheet]

        try:
            # First validate file structure / workbook layout
            errors, metadata = validate_structure(ws)
            if errors:
                raise F2000ValidationError(errors)

            # Next validate the content of each worksheet
            content_errors = validate_content(ws, metadata)
            if content_errors:
                raise F2000ValidationError(content_errors)

            # Resolve station from validated name
            station = find_station_by_name(metadata['station_name'])

            # Parse and insert
            reads.append( parse_data(ws, metadata, station, utc_offset) )
        except F2000ValidationError as e:
            logger.warning(f"Skipping sheet '{sheet}': {e}")
            continue
        except Exception as e:
            logger.warning(f"Skipping sheet '{sheet}': {repr(e)}")
            continue

    wb.close()

    try:
        insert(reads, override_data_on_conflict)
    except Exception as e:
        logger.error(repr(e))
        raise

    end = time.time()

    logger.info(f'Processing file {filename} in {end - start} seconds, '
                f'returning #reads={len(reads)}.')

    return reads
